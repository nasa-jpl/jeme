#!/usr/bin/env python3
"""
Generate TROPESS_sample_10_papers.xlsx — 10 randomly selected papers
with all available metadata columns.

Usage:
    python scripts/generate_sample_excel.py [--seed N] [--out PATH]
"""

import argparse
import json
import random
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_FILE = Path(__file__).parent.parent / "public" / "data" / "TROPESS_analyzed.json"
DEFAULT_OUT = Path(__file__).parent.parent / "TROPESS_sample_10_papers.xlsx"


def safe(val):
    """Flatten nested dicts/lists to a string safe for Excel."""
    if val is None:
        return ""
    if isinstance(val, dict):
        import json as _json
        return _json.dumps(val, ensure_ascii=False)
    if isinstance(val, list):
        parts = []
        for item in val:
            if isinstance(item, dict):
                parts.append(item.get("name") or item.get("title") or str(item))
            else:
                parts.append(str(item))
        return "; ".join(parts)
    return val


def flatten_paper(p):
    """Return an ordered dict of all columns for one paper."""
    uncertainty = p.get("uncertainty") or {}

    # Missions / instruments
    missions = p.get("missions_instruments") or []
    mission_names = "; ".join(
        m.get("name", "") if isinstance(m, dict) else str(m) for m in missions
    )

    # Authors
    authors = p.get("authors") or []
    if isinstance(authors, list):
        author_str = "; ".join(
            a.get("name", a) if isinstance(a, dict) else str(a) for a in authors
        )
    else:
        author_str = str(authors)

    return {
        # Identity
        "Title":                     safe(p.get("title")),
        "Authors":                   author_str,
        "Year":                      p.get("year", ""),
        "Venue":                     safe(p.get("venue")),
        "DOI":                       p.get("doi", ""),
        "URL":                       p.get("url", ""),
        "Paper ID":                  p.get("paper_id", ""),
        # Citation
        "Citation Count":            p.get("citation_count", 0),
        # Classification
        "Engagement Level":          p.get("engagement_level", ""),
        "Engagement Rationale":      p.get("engagement_level_rationale", ""),
        "Paper Type":                p.get("paper_type", ""),
        "Paper Type Rationale":      p.get("paper_type_rationale", ""),
        "Research Domain":           p.get("research_domain", ""),
        # Geography
        "Primary Country":           p.get("country", ""),
        "All Countries":             "; ".join(c for c in (p.get("all_countries") or [p.get("country")] or []) if c),
        # Team paper linkage
        "Citing Team Paper":         p.get("citing_team_paper", ""),
        "Team Paper ID":             p.get("team_paper_id", ""),
        # Peer review
        "Is Peer Reviewed":          str(p.get("is_peer_reviewed", "")),
        # Missions
        "Missions / Instruments":    mission_names,
        # Abstract
        "Abstract":                  p.get("abstract", "") or "",
        # Uncertainty block
        "Uncertainty Score":         uncertainty.get("composite_score", ""),
        "Evidence Confidence":       uncertainty.get("evidence_confidence", ""),
        "Reasoning Confidence":      uncertainty.get("reasoning_confidence", ""),
        "Pipeline Variance":         uncertainty.get("pipeline_variance", ""),
        "Miscalibration Risk":       uncertainty.get("miscalibration_risk", ""),
        "Stochastic Variance":       uncertainty.get("stochastic_variance", ""),
        "Override Flag":             str(uncertainty.get("override_flag", "")),
    }


HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
ALT_FILL     = PatternFill("solid", fgColor="EBF0FA")   # light blue-grey
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(size=9)
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
THIN_BORDER  = Border(
    left=Side(style="thin", color="D0D7E3"),
    right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)

# Column widths (col name → width chars)
COL_WIDTHS = {
    "Title": 55, "Authors": 35, "Year": 7, "Venue": 28, "DOI": 30,
    "URL": 30, "Paper ID": 38, "Citation Count": 10,
    "Engagement Level": 16, "Engagement Rationale": 55,
    "Paper Type": 12, "Paper Type Rationale": 55,
    "Research Domain": 30, "Primary Country": 18, "All Countries": 35,
    "Citing Team Paper": 55, "Team Paper ID": 38,
    "Is Peer Reviewed": 14, "Missions / Instruments": 35,
    "Abstract": 80,
    "Uncertainty Score": 14, "Evidence Confidence": 16,
    "Reasoning Confidence": 16, "Pipeline Variance": 14,
    "Miscalibration Risk": 16, "Stochastic Variance": 16, "Override Flag": 12,
}


def write_excel(rows, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TROPESS Sample Papers"

    headers = list(rows[0].keys())

    # Header row
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 28

    # Data rows
    for ri, row in enumerate(rows, start=2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=ri, column=ci, value=row[h])
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
        ws.row_dimensions[ri].height = 60

    # Column widths
    for ci, h in enumerate(headers, start=1):
        width = COL_WIDTHS.get(h, 20)
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Freeze header + auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(out_path)
    print(f"Saved: {out_path}  ({len(rows)} rows, {len(headers)} columns)")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--n", type=int, default=10)
    parser.add_argument("--out", default=str(DEFAULT_OUT))
    args = parser.parse_args()

    with open(DATA_FILE) as f:
        data = json.load(f)

    # Exclude team/seed papers (paper_id == a team_paper_id)
    team_paper_ids = {p.get('team_paper_id') for p in data if p.get('team_paper_id')}
    pool = [p for p in data if p.get('paper_id') not in team_paper_ids]
    print(f"Pool: {len(pool)} papers (excluded {len(data) - len(pool)} team papers)")

    random.seed(args.seed)
    sample = random.sample(pool, args.n)

    rows = [flatten_paper(p) for p in sample]
    write_excel(rows, args.out)

    # Quick summary
    print("\nSample classification summary:")
    for p in sample:
        print(f"  [{p.get('engagement_level','?'):16}] [{p.get('paper_type','?'):9}] "
              f"{p.get('title','')[:60]}")


if __name__ == "__main__":
    main()
